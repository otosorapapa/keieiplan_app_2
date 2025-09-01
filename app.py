
import streamlit as st
import pandas as pd
import numpy as np
import io
import math
import datetime as dt
from typing import Dict, Tuple, List, Any
import openpyxl  # noqa: F401  # Ensure Excel engine is available
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="経営計画策定（単年）｜Streamlit",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

DEFAULTS = {
    "sales": 1000000000,
    "fte": 20.0,
    "cogs_mat_rate": 0.25,
    "cogs_lbr_rate": 0.06,
    "cogs_out_src_rate": 0.1,
    "cogs_out_con_rate": 0.04,
    "cogs_oth_rate": 0.0,
    "opex_h_rate": 0.17,
    "opex_k_rate": 0.468,
    "opex_dep_rate": 0.006,
    "noi_misc_rate": 0.0001,
    "noi_grant_rate": 0.0,
    "noi_oth_rate": 0.0,
    "noe_int_rate": 0.0074,
    "noe_oth_rate": 0.0,
    "unit": "百万円",
    "fiscal_year": 2025
}

PLOT_STYLE_DEFAULT: Dict[str, Any] = {
    "figure_bg": "#FFFFFF",
    "axes_bg": "#FFFFFF",
    "grid": True,
    "grid_color": "#CCCCCC",
    "pos_color": "#1f77b4",
    "neg_color": "#d62728",
    "node_size": 10,
    "font_color": "#000000",
    "font_size": 10,
    "alpha": 0.9,
}

ITEMS = [
    ("REV", "売上高", "売上"),
    ("COGS_MAT", "外部仕入｜材料費", "外部仕入"),
    ("COGS_LBR", "外部仕入｜労務費(外部)", "外部仕入"),
    ("COGS_OUT_SRC", "外部仕入｜外注費(専属)", "外部仕入"),
    ("COGS_OUT_CON", "外部仕入｜外注費(委託)", "外部仕入"),
    ("COGS_OTH", "外部仕入｜その他諸経費", "外部仕入"),
    ("COGS_TTL", "外部仕入｜計", "外部仕入"),
    ("GROSS", "粗利(加工高)", "粗利"),
    ("OPEX_H", "内部費用｜人件費", "内部費用"),
    ("OPEX_K", "内部費用｜経費", "内部費用"),
    ("OPEX_DEP", "内部費用｜減価償却費", "内部費用"),
    ("OPEX_TTL", "内部費用｜計", "内部費用"),
    ("OP", "営業利益", "損益"),
    ("NOI_MISC", "営業外収益｜雑収入", "営業外"),
    ("NOI_GRANT", "営業外収益｜補助金/給付金", "営業外"),
    ("NOI_OTH", "営業外収益｜その他", "営業外"),
    ("NOE_INT", "営業外費用｜支払利息", "営業外"),
    ("NOE_OTH", "営業外費用｜雑損", "営業外"),
    ("ORD", "経常利益", "損益"),
    ("BE_SALES", "損益分岐点売上高", "KPI"),
    ("PC_SALES", "一人当たり売上", "KPI"),
    ("PC_GROSS", "一人当たり粗利", "KPI"),
    ("PC_ORD", "一人当たり経常利益", "KPI"),
    ("LDR", "労働分配率", "KPI")
]

# Mapping from item code to label for quick lookup
ITEM_LABELS = {code: label for code, label, _ in ITEMS}

# --- MCKINSEY TORNADO
def _set_jp_font() -> None:
    """日本語フォントを自動設定（環境に応じて存在チェック）"""
    for f in ["Yu Gothic", "Meiryo", "Hiragino Sans", "Noto Sans CJK JP", "IPAexGothic"]:
        try:
            mpl.font_manager.findfont(f, fallback_to_default=False)
            mpl.rcParams["font.family"] = f
            break
        except Exception:
            continue
    mpl.rcParams["axes.unicode_minus"] = False

def render_tornado_mckinsey(
    changes: List[Tuple[str, float]],
    title: str,
    unit_label: str,
    style: Dict[str, Any] | None = None,
) -> None:
    """マッキンゼー風トルネード図を描画しPNGダウンロードボタンを表示"""
    style = style or {}
    fig_bg = style.get("figure_bg", PLOT_STYLE_DEFAULT["figure_bg"])
    axes_bg = style.get("axes_bg", PLOT_STYLE_DEFAULT["axes_bg"])
    grid_on = style.get("grid", PLOT_STYLE_DEFAULT["grid"])
    grid_color = style.get("grid_color", PLOT_STYLE_DEFAULT["grid_color"])
    pos_color = style.get("pos_color", PLOT_STYLE_DEFAULT["pos_color"])
    neg_color = style.get("neg_color", PLOT_STYLE_DEFAULT["neg_color"])
    node_size = style.get("node_size", PLOT_STYLE_DEFAULT["node_size"])
    font_color = style.get("font_color", PLOT_STYLE_DEFAULT["font_color"])
    font_size = style.get("font_size", PLOT_STYLE_DEFAULT["font_size"])
    alpha = style.get("alpha", PLOT_STYLE_DEFAULT["alpha"])
    if not changes:
        st.warning("表示するデータがありません。")
        return
    changes_sorted = sorted(changes, key=lambda x: abs(x[1]), reverse=True)
    labels = [k for k, _ in changes_sorted]
    values = [v for _, v in changes_sorted]
    max_abs = max(abs(v) for v in values)
    if not math.isfinite(max_abs) or max_abs == 0:
        st.warning("有効なデータがありません。")
        return
    lim = max_abs * 1.1
    fig, ax = plt.subplots(figsize=(6, 0.45 * len(values) + 1))
    fig.patch.set_facecolor(fig_bg)
    ax.set_facecolor(axes_bg)
    y = np.arange(len(values))
    colors = [pos_color if v >= 0 else neg_color for v in values]
    bars = ax.barh(y, values, color=colors, alpha=alpha)
    ax.set_yticks(y, labels)
    ax.set_xlim(-lim, lim)
    ax.axvline(0, color=grid_color, linewidth=0.8)
    if grid_on:
        ax.grid(color=grid_color, axis="x", linewidth=0.5, linestyle="--")
    else:
        ax.grid(False)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(grid_color)
        ax.spines[spine].set_linewidth(0.5)
    ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f"¥{x:,.0f}"))
    ax.tick_params(axis="x", colors=font_color, labelsize=font_size)
    ax.tick_params(axis="y", colors=font_color, labelsize=font_size)
    ellipsis = False
    for bar, v in zip(bars, values):
        txt = f"{'+' if v >= 0 else '-'}¥{abs(v):,}"
        if abs(v) < lim * 0.05:
            txt = "..."
            ellipsis = True
        ax.text(
            v + (lim * 0.01 if v >= 0 else -lim * 0.01),
            bar.get_y() + bar.get_height() / 2,
            txt,
            ha="left" if v >= 0 else "right",
            va="center",
            clip_on=False,
            color=font_color,
            fontsize=node_size,
        )
    ax.set_title(title, color=font_color, fontsize=font_size + 2)
    fig.tight_layout()
    fig.text(0.5, -0.02, "注：右=利益増、左=利益減", ha="center", fontsize=font_size - 1, color=font_color)
    st.pyplot(fig, use_container_width=True)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
    st.download_button(
        "📥 感応度グラフ（PNG）",
        data=buf.getvalue(),
        file_name="tornado.png",
        mime="image/png",
    )
    if ellipsis:
        st.caption("※ 一部の値は省略記号で表示しています。下表で詳細を確認ください。")


def build_sensitivity_view_options():
    st.subheader("📉 感応度分析｜表示設定")
    c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
    with c1:
        viz = st.radio(
            "可視化タイプ",
            ["トルネード（±差分）", "ウォーターフォール（寄与累積）"],
            horizontal=True,
        )
    with c2:
        top_n = st.slider("表示項目数 (Top-N)", 3, 12, 6, 1)
    with c3:
        height_px = st.slider("グラフ高さ (px)", 200, 900, 360, 20)
    with c4:
        compact = st.checkbox("コンパクト表示（小さな文字）", True)

    step = st.slider("感応度ステップ（±）", 0.01, 0.20, 0.10, 0.01)
    show_values = st.checkbox("値ラベルを表示", True)
    return dict(viz=viz, top_n=top_n, height_px=height_px,
                compact=compact, step=step, show_values=show_values)


def _sensitivity_items(plan: dict, step: float):
    """各変数の±stepによる経常利益差分（ceteris paribus）。"""
    keys = [
        ("sales", "売上高", "amount"),
        ("gp_rate", "粗利率", "rate"),
        ("opex_h", "人件費", "amount"),
        ("opex_fixed", "販管費（固定費）", "amount"),
        ("opex_dep", "減価償却", "amount"),
        ("opex_oth", "その他費用", "amount"),
    ]
    base_ord = compute_plan(plan)["ord"]
    items = []
    for k, label, kind in keys:
        p_low = plan.copy()
        p_high = plan.copy()
        if kind == "rate":
            p_low[k] = max(0.0, plan[k] - step)
            p_high[k] = min(1.0, plan[k] + step)
        else:
            p_low[k] = max(0.0, plan[k] * (1 - step))
            p_high[k] = plan[k] * (1 + step)

        low_ord = compute_plan(p_low)["ord"]
        high_ord = compute_plan(p_high)["ord"]
        delta_low = low_ord - base_ord
        delta_high = high_ord - base_ord
        span = abs(delta_low) + abs(delta_high)
        items.append(dict(key=k, label=label,
                          delta_low=delta_low, delta_high=delta_high, span=span))
    items.sort(key=lambda x: x["span"], reverse=True)
    return items


def render_tornado_compact(plan: dict, step: float, top_n: int, height_px: int,
                           compact: bool, show_values: bool):
    """俯瞰性を高めたトルネード図（Top-N・高さ・フォント調整）"""
    items = _sensitivity_items(plan, step)[:top_n]
    labels = [x["label"] for x in items]
    lows = [x["delta_low"] for x in items]
    highs = [x["delta_high"] for x in items]

    fig_h_in = max(height_px / 96.0, 2 / 3)
    fig, ax = plt.subplots(figsize=(7, fig_h_in))
    for i, (lo, hi) in enumerate(zip(lows, highs)):
        ax.barh(i, hi, color="#0B3D91", alpha=0.9)
        ax.barh(i, lo, color="#9E9E9E", alpha=0.9)

    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels, fontsize=(9 if compact else 11))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f"¥{x:,.0f}"))
    ax.axvline(0, color="#D0D0D0", linewidth=0.8)
    ax.set_xlabel("経常利益への寄与（差分）", fontsize=(9 if compact else 11))

    if show_values:
        offset = max(1.0, max(abs(v) for v in lows + highs) * 0.02)
        for i, (lo, hi) in enumerate(zip(lows, highs)):
            ax.text(hi + (offset if hi >= 0 else -offset),
                    i, format_money(hi), va="center", ha="left" if hi >= 0 else "right", fontsize=(8 if compact else 10))
            ax.text(lo + (offset if lo >= 0 else -offset),
                    i, format_money(lo), va="center", ha="left" if lo >= 0 else "right", fontsize=(8 if compact else 10))

    fig.tight_layout()
    st.pyplot(fig, use_container_width=True)


def render_sensitivity_waterfall(plan: dict, step: float, top_n: int, height_px: int,
                                 compact: bool, show_values: bool):
    """
    感応度の「寄与累積」をウォーターフォールで表示。
    ・各変数を +step 側に単独シフトした場合の寄与を絶対値降順に並べ、
      ベースORDから順に累積表示（相互作用は考慮しない近似）。
    """
    base_ord = compute_plan(plan)["ord"]
    items = _sensitivity_items(plan, step)[:top_n]
    contribs = [(x["label"], x["delta_high"]) for x in items]
    labels = ["ベースORD"] + [lbl for lbl, _ in contribs] + ["概算ORD（+step適用）"]
    vals = [base_ord] + [v for _, v in contribs] + [0.0]
    cum = [vals[0]]
    for v in vals[1:-1]:
        cum.append(cum[-1] + v)
    final = cum[-1]
    vals[-1] = final - base_ord

    fig_h_in = max(height_px / 96.0, 2 / 3)
    fig, ax = plt.subplots(figsize=(7, fig_h_in))
    colors = []
    for i, v in enumerate(vals):
        if i == 0 or i == len(vals) - 1:
            colors.append("#0B3D91")
        else:
            colors.append("#0B3D91" if v >= 0 else "#9E9E9E")

    ax.bar(range(len(vals)), vals, color=colors)
    ax.axhline(0, color="#D0D0D0", linewidth=0.8)
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=(8 if compact else 10))
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"¥{x:,.0f}"))
    ax.set_ylabel("寄与（累積）", fontsize=(9 if compact else 11))

    if show_values:
        ref = max(1.0, max(abs(v) for v in vals))
        for i, v in enumerate(vals):
            ax.text(i, v + (0.02 * ref if v >= 0 else -0.02 * ref),
                    format_money(v), ha="center",
                    va="bottom" if v >= 0 else "top",
                    fontsize=(8 if compact else 10))

    fig.tight_layout()
    st.pyplot(fig, use_container_width=True)


def render_sensitivity_view(plan: dict):
    """感応度分析ビューの統括（俯瞰性改善＋ウォーターフォール追加）"""
    opt = build_sensitivity_view_options()
    if opt["viz"].startswith("トルネード"):
        render_tornado_compact(plan, opt["step"], opt["top_n"], opt["height_px"],
                               opt["compact"], opt["show_values"])
    else:
        render_sensitivity_waterfall(plan, opt["step"], opt["top_n"], opt["height_px"],
                                     opt["compact"], opt["show_values"])

# --- EXCEL JP LOCALE
def apply_japanese_styles(wb) -> None:
    """ヘッダ太字・中央揃え、列幅自動調整、1行目固定"""
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(28, max(10, max_len + 2))

def format_money_and_percent(ws, money_cols: List[int], percent_cols: List[int]) -> None:
    """通貨および百分率の書式を適用"""
    money_fmt = "\"¥\"#,##0;[Red]-\"¥\"#,##0"
    for c in money_cols:
        col_letter = get_column_letter(c)
        for cell in ws[col_letter][1:]:
            cell.number_format = money_fmt
    for c in percent_cols:
        col_letter = get_column_letter(c)
        for cell in ws[col_letter][1:]:
            cell.number_format = "0.0%"

def millions(x):
    return x / 1_000_000

def thousands(x):
    return x / 1_000

def format_money(x, unit="百万円"):
    if x is None or (isinstance(x, float) and (np.isnan(x) or np.isinf(x))):
        return "—"
    if unit == "百万円":
        return f"{millions(x):,.1f}"
    elif unit == "千円":
        return f"{thousands(x):,.0f}"
    else:
        return f"{x:,.0f}"


def compute_plan(plan: dict) -> dict:
    """Aggregate plan calculation returning ordinary profit.

    Parameters
    ----------
    plan: dict
        Dictionary containing at minimum the keys ``sales``, ``gp_rate``,
        ``opex_h``, ``opex_fixed``, ``opex_dep`` and ``opex_oth``.

    Returns
    -------
    dict
        A dictionary with the key ``ord`` representing ordinary profit.
    """
    sales = float(plan.get("sales", 0.0))
    gp_rate = float(plan.get("gp_rate", 0.0))
    gross = sales * gp_rate
    opex_h = float(plan.get("opex_h", 0.0))
    opex_fixed = float(plan.get("opex_fixed", 0.0))
    opex_dep = float(plan.get("opex_dep", 0.0))
    opex_oth = float(plan.get("opex_oth", 0.0))
    ord_profit = gross - opex_h - opex_fixed - opex_dep - opex_oth
    return {"ord": ord_profit}

class PlanConfig:
    def __init__(self, base_sales: float, fte: float, unit: str) -> None:
        self.base_sales = base_sales
        self.fte = max(0.0001, fte)
        self.unit = unit
        self.items: Dict[str, Dict[str, float]] = {}

    def set_rate(self, code: str, rate: float, rate_base: str = 'sales') -> None:
        self.items[code] = {"method": "rate", "value": float(rate), "rate_base": rate_base}

    def set_amount(self, code: str, amount: float) -> None:
        self.items[code] = {"method": "amount", "value": float(amount), "rate_base": "fixed"}

    def clone(self) -> 'PlanConfig':
        c = PlanConfig(self.base_sales, self.fte, self.unit)
        c.items = {k: v.copy() for k, v in self.items.items()}
        return c


def dual_input_row(label: str, base_sales: float, *,
                   mode_key: str,
                   pct_default: float = 0.0,
                   amount_default: float = 0.0,
                   pct_min: float = 0.0, pct_max: float = 3.0, pct_step: float = 0.005,
                   help_text: str = "") -> dict:
    """
    返り値: {"method": "rate" or "amount", "value": float}
    - mode=="％（増減/売上対比）": 率を編集、実額は参考表示（= rate * base_sales）
    - mode=="実額（円）": 実額を編集、率は参考表示（= amount / base_sales）
    - 0除算/NaNは自動で保護し、表示は0とする
    """
    mode = st.session_state.get(mode_key, "％（増減/売上対比）")
    key_base = label.replace("｜", "_").replace(" ", "_")
    if mode == "％（増減/売上対比）":
        rate = st.number_input(
            f"{label}（率）",
            min_value=pct_min,
            max_value=pct_max,
            step=pct_step,
            format="%.3f",
            value=pct_default,
            help=help_text,
            key=f"{key_base}_pct"
        )
        amount = rate * base_sales
        if not math.isfinite(amount):
            amount = 0.0
        st.caption(f"金額 ¥{amount:,.0f}")
        return {"method": "rate", "value": rate}
    else:
        amount = st.number_input(
            f"{label}（実額）",
            min_value=0.0,
            step=1_000_000.0,
            format="%.0f",
            value=amount_default,
            help=help_text,
            key=f"{key_base}_amt"
        )
        if not math.isfinite(amount):
            amount = 0.0
        rate = amount / base_sales if base_sales > 0 else 0.0
        if not math.isfinite(rate):
            rate = 0.0
        st.caption(f"率 {rate*100:.1f}%")
        return {"method": "amount", "value": amount}

def compute(plan: PlanConfig, sales_override: float | None = None, amount_overrides: Dict[str, float] | None = None) -> Dict[str, float]:
    S = float(plan.base_sales if sales_override is None else sales_override)
    amt = {code: 0.0 for code, *_ in ITEMS}
    amt["REV"] = S

    def line_amount(code, gross_guess):
        cfg = plan.items.get(code, None)
        if amount_overrides and code in amount_overrides:
            return float(amount_overrides[code])
        if cfg is None:
            return 0.0
        if cfg["method"] == "amount":
            return float(cfg["value"])
        r = float(cfg["value"])
        base = cfg.get("rate_base", "sales")
        if base == "sales":
            return S * r
        elif base == "gross":
            return max(0.0, gross_guess) * r
        elif base == "fixed":
            return r
        return S * r

    cogs_codes = ["COGS_MAT", "COGS_LBR", "COGS_OUT_SRC", "COGS_OUT_CON", "COGS_OTH"]
    sales_based_cogs = 0.0
    for code in cogs_codes:
        cfg = plan.items.get(code)
        if cfg and cfg["method"] == "rate" and cfg.get("rate_base", "sales") == "sales":
            sales_based_cogs += S * float(cfg["value"])
        elif cfg and cfg["method"] == "amount":
            sales_based_cogs += float(cfg["value"])

    gross = S - sales_based_cogs
    for _ in range(5):
        cogs = 0.0
        for code in cogs_codes:
            cogs += max(0.0, line_amount(code, gross))
        gross_new = S - cogs
        if abs(gross_new - gross) < 1e-6:
            gross = gross_new
            break
        gross = gross_new

    cogs_total = 0.0
    for code in cogs_codes:
        val = max(0.0, line_amount(code, gross))
        amt[code] = val
        cogs_total += val
    amt["COGS_TTL"] = cogs_total
    amt["GROSS"] = S - cogs_total

    opex_codes = ["OPEX_H", "OPEX_K", "OPEX_DEP"]
    opex_total = 0.0
    for code in opex_codes:
        val = max(0.0, line_amount(code, amt["GROSS"]))
        amt[code] = val
        opex_total += val
    amt["OPEX_TTL"] = opex_total

    amt["OP"] = amt["GROSS"] - amt["OPEX_TTL"]

    noi_codes = ["NOI_MISC", "NOI_GRANT", "NOI_OTH"]
    noe_codes = ["NOE_INT", "NOE_OTH"]
    for code in noi_codes + noe_codes:
        val = max(0.0, line_amount(code, amt["GROSS"]))
        amt[code] = val

    amt["ORD"] = amt["OP"] + (amt["NOI_MISC"] + amt["NOI_GRANT"] + amt["NOI_OTH"]) - (amt["NOE_INT"] + amt["NOE_OTH"])

    var_cost = 0.0
    for code in cogs_codes + opex_codes + noi_codes + noe_codes:
        cfg = plan.items.get(code)
        if cfg and cfg["method"] == "rate" and cfg.get("rate_base", "sales") in ("sales", "gross"):
            if cfg.get("rate_base") == "gross":
                g_ratio = amt["GROSS"] / S if S > 0 else 0.0
                var_cost += S * (cfg["value"] * g_ratio)
            else:
                var_cost += S * cfg["value"]

    fixed_cost = 0.0
    for code in cogs_codes + opex_codes + noi_codes + noe_codes:
        cfg = plan.items.get(code)
        if cfg and cfg["method"] == "amount":
            fixed_cost += cfg["value"]
        elif cfg and cfg.get("rate_base") == "fixed":
            if cfg["method"] == "rate":
                fixed_cost += cfg["value"]
    cm_ratio = 1.0 - (var_cost / S if S > 0 else 0.0)
    if cm_ratio <= 0:
        be_sales = float("inf")
    else:
        be_sales = fixed_cost / cm_ratio
    amt["BE_SALES"] = be_sales

    fte = max(0.0001, plan.fte)
    amt["PC_SALES"] = amt["REV"] / fte
    amt["PC_GROSS"] = amt["GROSS"] / fte
    amt["PC_ORD"] = amt["ORD"] / fte
    amt["LDR"] = (amt["OPEX_H"] / amt["GROSS"]) if amt["GROSS"] > 0 else np.nan

    return amt

def bisection_for_target_op(plan: PlanConfig, target_op: float, s_low: float, s_high: float, max_iter=60, eps=1_000.0) -> Tuple[float, Dict[str, float]]:
    def op_at(S):
        return compute(plan, sales_override=S)["ORD"]
    low, high = max(0.0, s_low), max(s_low * 1.5, s_high)
    f_low = op_at(low)
    f_high = op_at(high)
    it = 0
    while (f_low - target_op) * (f_high - target_op) > 0 and high < 1e13 and it < 40:
        high = high * 1.6 if high > 0 else 1_000_000.0
        f_high = op_at(high)
        it += 1
    for _ in range(max_iter):
        mid = 0.5 * (low + high)
        f_mid = op_at(mid)
        if abs(f_mid - target_op) <= eps:
            return mid, compute(plan, sales_override=mid)
        if (f_low - target_op) * (f_mid - target_op) <= 0:
            high, f_high = mid, f_mid
        else:
            low, f_low = mid, f_mid
    mid = 0.5 * (low + high)
    return mid, compute(plan, sales_override=mid)

# Sidebar
mode = st.sidebar.radio(
    "入力モード",
    ["％（増減/売上対比）", "実額（円）"],
    horizontal=True,
    index=0,
    key="input_mode",
)

with st.sidebar:
    st.header("⚙️ 基本設定")
    fiscal_year = st.number_input("会計年度", value=int(DEFAULTS["fiscal_year"]), step=1, format="%d")
    unit = st.selectbox("表示単位", ["百万円", "千円", "円"], index=0, help="計算は円ベース、表示のみ丸めます。")
    base_sales = st.number_input("売上高（ベース）", value=float(DEFAULTS["sales"]), step=10_000_000.0, min_value=0.0, format="%.0f")
    fte = st.number_input("人員数（FTE換算）", value=float(DEFAULTS["fte"]), step=1.0, min_value=0.0)

    st.markdown("---")
    st.caption("外部仕入")
    cogs_mat_input = dual_input_row(
        "材料費",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["cogs_mat_rate"]),
        amount_default=base_sales * DEFAULTS["cogs_mat_rate"],
        pct_step=0.01,
    )
    cogs_lbr_input = dual_input_row(
        "労務費(外部)",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["cogs_lbr_rate"]),
        amount_default=base_sales * DEFAULTS["cogs_lbr_rate"],
        pct_step=0.01,
    )
    cogs_out_src_input = dual_input_row(
        "外注費(専属)",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["cogs_out_src_rate"]),
        amount_default=base_sales * DEFAULTS["cogs_out_src_rate"],
        pct_step=0.01,
    )
    cogs_out_con_input = dual_input_row(
        "外注費(委託)",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["cogs_out_con_rate"]),
        amount_default=base_sales * DEFAULTS["cogs_out_con_rate"],
        pct_step=0.01,
    )
    cogs_oth_input = dual_input_row(
        "その他諸経費",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["cogs_oth_rate"]),
        amount_default=base_sales * DEFAULTS["cogs_oth_rate"],
        pct_step=0.005,
    )

    st.markdown("---")
    st.caption("内部費用")
    opex_h_input = dual_input_row(
        "人件費",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["opex_h_rate"]),
        amount_default=base_sales * DEFAULTS["opex_h_rate"],
        pct_step=0.01,
    )
    opex_k_input = dual_input_row(
        "経費",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["opex_k_rate"]),
        amount_default=base_sales * DEFAULTS["opex_k_rate"],
        pct_step=0.01,
    )
    opex_dep_input = dual_input_row(
        "減価償却",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["opex_dep_rate"]),
        amount_default=base_sales * DEFAULTS["opex_dep_rate"],
        pct_step=0.001,
    )

    st.markdown("---")
    st.caption("営業外")
    noi_misc_input = dual_input_row(
        "営業外収益：雑収入",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["noi_misc_rate"]),
        amount_default=base_sales * DEFAULTS["noi_misc_rate"],
        pct_min=0.0,
        pct_max=1.0,
        pct_step=0.0005,
    )
    noi_grant_input = dual_input_row(
        "営業外収益：補助金",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["noi_grant_rate"]),
        amount_default=base_sales * DEFAULTS["noi_grant_rate"],
        pct_min=0.0,
        pct_max=1.0,
        pct_step=0.0005,
    )
    noi_oth_input = dual_input_row(
        "営業外収益：その他",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["noi_oth_rate"]),
        amount_default=base_sales * DEFAULTS["noi_oth_rate"],
        pct_min=0.0,
        pct_max=1.0,
        pct_step=0.0005,
    )
    noe_int_input = dual_input_row(
        "営業外費用：支払利息",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["noe_int_rate"]),
        amount_default=base_sales * DEFAULTS["noe_int_rate"],
        pct_min=0.0,
        pct_max=1.0,
        pct_step=0.0005,
    )
    noe_oth_input = dual_input_row(
        "営業外費用：雑損",
        base_sales,
        mode_key="input_mode",
        pct_default=float(DEFAULTS["noe_oth_rate"]),
        amount_default=base_sales * DEFAULTS["noe_oth_rate"],
        pct_min=0.0,
        pct_max=1.0,
        pct_step=0.0005,
    )

    st.markdown("---")
    st.header("🎨 グラフスタイル")
    fig_bg = st.color_picker("図背景色", PLOT_STYLE_DEFAULT["figure_bg"])
    axes_bg = st.color_picker("枠背景色", PLOT_STYLE_DEFAULT["axes_bg"])
    show_grid = st.checkbox("グリッド線を表示", value=PLOT_STYLE_DEFAULT["grid"])
    grid_color = st.color_picker("グリッド線色", PLOT_STYLE_DEFAULT["grid_color"])
    pos_color = st.color_picker("増加色", PLOT_STYLE_DEFAULT["pos_color"])
    neg_color = st.color_picker("減少色", PLOT_STYLE_DEFAULT["neg_color"])
    node_size = st.slider("ノードサイズ", 1, 30, PLOT_STYLE_DEFAULT["node_size"])
    font_color = st.color_picker("フォント色", PLOT_STYLE_DEFAULT["font_color"])
    font_size = st.slider("フォントサイズ", 6, 24, PLOT_STYLE_DEFAULT["font_size"])
    alpha = st.slider("透過度", 0.0, 1.0, PLOT_STYLE_DEFAULT["alpha"], 0.05)

plot_style = {
    "figure_bg": fig_bg,
    "axes_bg": axes_bg,
    "grid": show_grid,
    "grid_color": grid_color,
    "pos_color": pos_color,
    "neg_color": neg_color,
    "node_size": node_size,
    "font_color": font_color,
    "font_size": font_size,
    "alpha": alpha,
}

base_plan = PlanConfig(base_sales=base_sales, fte=fte, unit=unit)


def apply_setting(code: str, result: dict) -> None:
    if result["method"] == "rate":
        base_plan.set_rate(code, result["value"], "sales")
    else:
        base_plan.set_amount(code, result["value"])


apply_setting("COGS_MAT", cogs_mat_input)
apply_setting("COGS_LBR", cogs_lbr_input)
apply_setting("COGS_OUT_SRC", cogs_out_src_input)
apply_setting("COGS_OUT_CON", cogs_out_con_input)
apply_setting("COGS_OTH", cogs_oth_input)

apply_setting("OPEX_H", opex_h_input)
apply_setting("OPEX_K", opex_k_input)
apply_setting("OPEX_DEP", opex_dep_input)

apply_setting("NOI_MISC", noi_misc_input)
apply_setting("NOI_GRANT", noi_grant_input)
apply_setting("NOI_OTH", noi_oth_input)
apply_setting("NOE_INT", noe_int_input)
apply_setting("NOE_OTH", noe_oth_input)

tab_input, tab_scen, tab_analysis, tab_export = st.tabs(["📝 計画入力", "🧪 シナリオ", "📊 感応度分析", "📤 エクスポート"])

with tab_input:
    st.subheader("単年利益計画（目標列）")
    base_amt = compute(base_plan)
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("売上高", f"{format_money(base_amt['REV'], base_plan.unit)} {base_plan.unit}")
    c2.metric("粗利(加工高)", f"{format_money(base_amt['GROSS'], base_plan.unit)} {base_plan.unit}")
    c3.metric("営業利益", f"{format_money(base_amt['OP'], base_plan.unit)} {base_plan.unit}")
    c4.metric("経常利益", f"{format_money(base_amt['ORD'], base_plan.unit)} {base_plan.unit}")
    be_label = "∞" if not math.isfinite(base_amt["BE_SALES"]) else f"{format_money(base_amt['BE_SALES'], base_plan.unit)} {base_plan.unit}"
    c5.metric("損益分岐点売上高", be_label)

    c6, c7, c8 = st.columns(3)
    c6.metric("一人当たり売上", f"{format_money(base_amt['PC_SALES'], base_plan.unit)} {base_plan.unit}")
    c7.metric("一人当たり粗利", f"{format_money(base_amt['PC_GROSS'], base_plan.unit)} {base_plan.unit}")
    ldr = base_amt["LDR"]
    ldr_str = "—" if (ldr is None or not math.isfinite(ldr)) else f"{ldr*100:.1f}%"
    c8.metric("労働分配率", ldr_str)

    rows = []
    for code, label, group in ITEMS:
        if code in ("PC_SALES", "PC_GROSS", "PC_ORD", "LDR", "BE_SALES"):
            continue
        val = base_amt[code]
        rows.append({"項目": label, "金額": format_money(val, base_plan.unit)})
    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, height=min(520, 40 + 28*len(rows)))

    st.info("ヒント: サイドバーの％／実額・人員・売上を変えると、即座に計算結果が更新されます。さらに固定費や個別額を設定したい場合は、下の『金額上書き』を利用してください。")

    with st.expander("🔧 金額上書き（固定費/個別額の設定）", expanded=False):
        st.caption("金額が入力された項目は、率の指定より優先され固定費扱いになります。")
        col1, col2, col3 = st.columns(3)
        override_inputs = {}
        for i, code in enumerate(["COGS_MAT","COGS_LBR","COGS_OUT_SRC","COGS_OUT_CON","COGS_OTH","OPEX_H","OPEX_K","OPEX_DEP","NOI_MISC","NOI_GRANT","NOI_OTH","NOE_INT","NOE_OTH"]):
            if i % 3 == 0:
                c = col1
            elif i % 3 == 1:
                c = col2
            else:
                c = col3
            # Look up label without reconstructing the dictionary each time
            val = c.number_input(
                f"{ITEM_LABELS[code]}（金額上書き）",
                min_value=0.0,
                value=0.0,
                step=1_000_000.0,
                key=f"ov_{code}"
            )
            if val > 0:
                override_inputs[code] = val

        if st.button("上書きを反映", type="primary"):
            preview_amt = compute(base_plan, amount_overrides=override_inputs)
            st.session_state["overrides"] = override_inputs
            st.success("上書きを反映しました（この状態でシナリオにも適用されます）。")

            rows2 = []
            for code, label, group in ITEMS:
                if code in ("PC_SALES","PC_GROSS","PC_ORD","LDR","BE_SALES"):
                    continue
                before = base_amt[code]
                after = preview_amt[code]
                rows2.append({"項目": label, "前": format_money(before, base_plan.unit), "後": format_money(after, base_plan.unit)})
            st.dataframe(pd.DataFrame(rows2), use_container_width=True)

def scenario_table(plan: PlanConfig, unit: str, overrides: Dict[str, float]) -> Tuple[pd.DataFrame, pd.DataFrame, List[Tuple[str, Dict[str, float]]]]:
    # --- SCENARIO UX
    type_display = ["なし", "売上高±%", "粗利率±pt", "目標経常", "昨年同一", "BEP"]
    type_map = {"なし": "none", "売上高±%": "sales_pct", "粗利率±pt": "gross_pt", "目標経常": "target_op", "昨年同一": "last_year", "BEP": "bep"}
    default_specs = [
        {"名称": "目標", "タイプ": "なし", "値": None},
        {"名称": "売上高10%増", "タイプ": "売上高±%", "値": 10.0},
        {"名称": "売上高5%減", "タイプ": "売上高±%", "値": -5.0},
        {"名称": "売上高10%減", "タイプ": "売上高±%", "値": -10.0},
        {"名称": "粗利1%減", "タイプ": "粗利率±pt", "値": -1.0},
        {"名称": "経常利益5千万円", "タイプ": "目標経常", "値": 50_000_000.0},
        {"名称": "昨年同一", "タイプ": "昨年同一", "値": None},
        {"名称": "損益分岐点売上高", "タイプ": "BEP", "値": None},
    ]
    df = st.session_state.get("scenario_df")
    if df is None:
        df = pd.DataFrame(default_specs)
    st.caption("各シナリオのラベルとパラメータを編集できます。")
    editor = st.data_editor(
        df,
        key="scenario_editor",
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "名称": st.column_config.TextColumn("名称"),
            "タイプ": st.column_config.SelectboxColumn("タイプ", options=type_display),
            "値": st.column_config.NumberColumn("値", help="タイプにより入力範囲が異なります"),
        },
    )
    st.session_state["scenario_df"] = editor.copy()

    def apply_driver(plan: PlanConfig, spec: Dict[str, float], overrides_local: Dict[str, float]):
        t = spec["type"]
        v = spec.get("value", None)
        if t == "none":
            return plan.base_sales, overrides_local, None
        if t == "sales_pct":
            S = plan.base_sales * (1.0 + float(v))
            return S, overrides_local, None
        if t == "gross_pt":
            delta = float(v)
            S = plan.base_sales
            delta_e = -delta * S
            ov = dict(overrides_local) if overrides_local else {}
            current = ov.get("COGS_OTH", None)
            if current is None:
                tmp = compute(plan, sales_override=S, amount_overrides=ov)
                base_oth = tmp["COGS_OTH"]
                ov["COGS_OTH"] = max(0.0, base_oth + delta_e)
            else:
                ov["COGS_OTH"] = max(0.0, current + delta_e)
            return S, ov, None
        if t == "target_op":
            target = float(v)
            sol_S, sol_amt = bisection_for_target_op(plan, target, s_low=0.0, s_high=max(1.2 * plan.base_sales, 1_000_000.0))
            return sol_S, overrides_local, sol_amt
        if t == "last_year":
            return plan.base_sales, overrides_local, None
        if t == "bep":
            temp = compute(plan, sales_override=plan.base_sales, amount_overrides=overrides_local)
            be = temp["BE_SALES"]
            return be if math.isfinite(be) else plan.base_sales, overrides_local, None
        return plan.base_sales, overrides_local, None

    b1, b2, b3, b4, b5 = st.columns(5)
    if b1.button("➕ 追加"):
        new_name = f"シナリオ{len(editor)+1}"
        editor.loc[len(editor)] = [new_name, "なし", None]
        st.session_state["scenario_df"] = editor
    if b2.button("🗑️ 選択行を削除"):
        sel = st.session_state.get("scenario_editor", {}).get("selected_rows", [])
        if sel:
            editor = editor.drop(index=sel).reset_index(drop=True)
            st.session_state["scenario_df"] = editor
    if b3.button("⟳ 既定にリセット"):
        editor = pd.DataFrame(default_specs)
        st.session_state["scenario_df"] = editor
    if b4.button("📌 保存"):
        st.session_state["scenarios"] = editor.to_dict(orient="records")
        st.success("保存しました。")
    if b5.button("📥 読込") and "scenarios" in st.session_state:
        editor = pd.DataFrame(st.session_state["scenarios"])
        st.session_state["scenario_df"] = editor

    selected = st.session_state.get("scenario_editor", {}).get("selected_rows", [])
    if len(selected) == 1:
        idx = selected[0]
        row = editor.loc[idx]
        typ_code = type_map.get(row["タイプ"], "none")
        with st.expander(f"詳細設定：{row['名称']}", expanded=True):
            if typ_code == "sales_pct":
                val = st.slider("売上高±%", -50.0, 50.0, float(row["値"] or 0.0), 1.0)
                editor.at[idx, "値"] = val
            elif typ_code == "gross_pt":
                val = st.slider("粗利率±pt", -10.0, 10.0, float(row["値"] or 0.0), 0.5, help="1pt=1%ポイント")
                editor.at[idx, "値"] = val
            elif typ_code == "target_op":
                val = st.number_input("目標経常利益（円）", min_value=0.0, value=float(row["値"] or 0.0), step=1_000_000.0, format="%.0f")
                editor.at[idx, "値"] = val
            else:
                st.write("—")
        st.session_state["scenario_df"] = editor
        spec = {"type": typ_code, "value": editor.at[idx, "値"]}
        base_amt = compute(plan, amount_overrides=overrides)
        S_override, ov, pre_amt = apply_driver(plan, spec, overrides)
        amt_prev = compute(plan, sales_override=S_override, amount_overrides=ov) if pre_amt is None else pre_amt
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("REV", f"{format_money(amt_prev['REV'], unit)} {unit}")
        c2.metric("GROSS", f"{format_money(amt_prev['GROSS'], unit)} {unit}")
        c3.metric("ORD", f"{format_money(amt_prev['ORD'], unit)} {unit}")
        be_lbl = "∞" if not math.isfinite(amt_prev['BE_SALES']) else f"{format_money(amt_prev['BE_SALES'], unit)} {unit}"
        c4.metric("BE_SALES", be_lbl)

    editable = []
    for _, row in editor.iterrows():
        typ_code = type_map.get(row["タイプ"], "none")
        val = row["値"]
        val = None if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))) else float(val)
        editable.append((row["名称"], {"type": typ_code, "value": val}))

    cols = ["項目"] + [nm for nm, _ in editable]
    rows = {code: [label] for code, label, _ in ITEMS if code not in ("PC_SALES", "PC_GROSS", "PC_ORD", "LDR", "BE_SALES")}
    kpis = {"BE_SALES": ["損益分岐点売上高"], "PC_SALES": ["一人当たり売上"], "PC_GROSS": ["一人当たり粗利"], "PC_ORD": ["一人当たり経常利益"], "LDR": ["労働分配率"]}

    base_amt = compute(plan, amount_overrides=overrides)
    for code, label, _ in ITEMS:
        if code in rows:
            rows[code].append(format_money(base_amt.get(code, 0.0), unit))
    for k in kpis.keys():
        if k == "LDR":
            val = base_amt.get("LDR", float("nan"))
            kpis[k].append(f"{val*100:.1f}%" if val == val else "—")
        else:
            kpis[k].append(format_money(base_amt.get(k, 0.0), unit))

    for nm, spec in editable[1:]:
        S_override, ov, pre_amt = apply_driver(plan, spec, overrides)
        scn_amt = compute(plan, sales_override=S_override, amount_overrides=ov) if pre_amt is None else pre_amt
        for code, label, _ in ITEMS:
            if code in rows:
                rows[code].append(format_money(scn_amt.get(code, 0.0), unit))
        for k in kpis.keys():
            if k == "LDR":
                v = scn_amt.get("LDR", float("nan"))
                kpis[k].append(f"{v*100:.1f}%" if v == v else "—")
            else:
                kpis[k].append(format_money(scn_amt.get(k, 0.0), unit))

    df1 = pd.DataFrame(rows.values(), columns=cols, index=rows.keys())
    df2 = pd.DataFrame(kpis.values(), columns=cols, index=kpis.keys())
    st.subheader("シナリオ比較（金額）")
    st.dataframe(df1, use_container_width=True)
    st.subheader("KPI（損益分岐点・一人当たり・労働分配率）")
    st.dataframe(df2, use_container_width=True)
    return df1, df2, editable

with tab_scen:
    overrides = st.session_state.get("overrides", {})
    df_amounts, df_kpis, scenario_specs = scenario_table(base_plan, unit, overrides)

with tab_analysis:
    _set_jp_font()
    base_amt = compute(base_plan, amount_overrides=st.session_state.get("overrides", {}))
    plan_inputs = {
        "sales": base_amt["REV"],
        "gp_rate": (base_amt["GROSS"] / base_amt["REV"]) if base_amt["REV"] else 0.0,
        "opex_h": base_amt["OPEX_H"],
        "opex_fixed": base_amt["OPEX_K"],
        "opex_dep": base_amt["OPEX_DEP"],
        "opex_oth": -(base_amt["NOI_MISC"] + base_amt["NOI_GRANT"] + base_amt["NOI_OTH"]
                       - base_amt["NOE_INT"] - base_amt["NOE_OTH"]),
    }
    render_sensitivity_view(plan_inputs)

with tab_export:
    st.subheader("エクスポート")
    st.caption("ワンクリックでExcel出力（シート: 金額, KPI, 感応度）。PDFはExcelから印刷設定で作成してください。")
    specs = scenario_specs

    def compute_scenario_numeric(plan, specs, overrides):
        cols = ["項目"] + [nm for nm,_ in specs]
        num_rows = {code: [label] for code, label, _ in ITEMS if code not in ("PC_SALES","PC_GROSS","PC_ORD","LDR","BE_SALES")}
        num_kpis = {"BE_SALES": ["損益分岐点売上高"], "PC_SALES": ["一人当たり売上"], "PC_GROSS": ["一人当たり粗利"], "PC_ORD": ["一人当たり経常利益"], "LDR": ["労働分配率"]}
        def apply_driver(spec):
            t = spec["type"]; v = spec.get("value", None)
            if t == "none": return plan.base_sales, overrides, None
            if t == "sales_pct": return plan.base_sales * (1.0 + float(v)), overrides, None
            if t == "gross_pt":
                S = plan.base_sales
                delta_e = -float(v) * S
                ov = dict(overrides) if overrides else {}
                tmp = compute(plan, sales_override=S, amount_overrides=ov)
                base_oth = tmp["COGS_OTH"]
                ov["COGS_OTH"] = max(0.0, base_oth + delta_e)
                return S, ov, None
            if t == "target_op":
                target = float(v)
                sol_S, sol_amt = bisection_for_target_op(plan, target, s_low=0.0, s_high=max(1.2*plan.base_sales, 1_000_000.0))
                return sol_S, overrides, sol_amt
            if t == "last_year":
                return plan.base_sales, overrides, None
            if t == "bep":
                temp = compute(plan, sales_override=plan.base_sales, amount_overrides=overrides)
                be = temp["BE_SALES"]
                return be if math.isfinite(be) else plan.base_sales, overrides, None
            return plan.base_sales, overrides, None

        base_amt = compute(plan, amount_overrides=overrides)
        for code, label, _ in ITEMS:
            if code in num_rows:
                num_rows[code].append(base_amt.get(code, 0.0))
        for k in num_kpis.keys():
            num_kpis[k].append(base_amt.get(k, 0.0))

        for (nm, spec) in specs[1:]:
            S, ov, pre = apply_driver(spec)
            scn_amt = compute(plan, sales_override=S, amount_overrides=ov) if pre is None else pre
            for code, label, _ in ITEMS:
                if code in num_rows:
                    num_rows[code].append(scn_amt.get(code, 0.0))
            for k in num_kpis.keys():
                num_kpis[k].append(scn_amt.get(k, 0.0))

        df_num = pd.DataFrame(num_rows.values(), columns=cols, index=num_rows.keys())
        df_kpi = pd.DataFrame(num_kpis.values(), columns=cols, index=num_kpis.keys())
        return df_num, df_kpi

    df_num, df_kpi = compute_scenario_numeric(base_plan, specs, st.session_state.get("overrides", {}))

    def recompute_sensitivity_table():
        base_amt = compute(base_plan, amount_overrides=st.session_state.get("overrides", {}))
        base_op = base_amt["ORD"]
        def op_with(ds=0.1, dgp=0.01, dH=0.1, dK=0.1):
            plan = base_plan.clone()
            S = plan.base_sales * (1.0 + ds)
            overrides = st.session_state.get("overrides", {}).copy()
            delta_e = -dgp * S
            overrides["COGS_OTH"] = max(0.0, compute(plan, sales_override=S, amount_overrides=overrides)["COGS_OTH"] + delta_e)
            val = compute(plan, sales_override=S, amount_overrides=overrides)["OPEX_H"]
            overrides["OPEX_H"] = max(0.0, val * (1.0 + dH))
            val = compute(plan, sales_override=S, amount_overrides=overrides)["OPEX_K"]
            overrides["OPEX_K"] = max(0.0, val * (1.0 + dK))
            return compute(plan, sales_override=S, amount_overrides=overrides)["ORD"]
        changes = [
            ("売上高 +10%", op_with(ds=+0.10) - base_op),
            ("売上高 -10%", op_with(ds=-0.10) - base_op),
            ("粗利率 +1pt", op_with(dgp=+0.01) - base_op),
            ("粗利率 -1pt", op_with(dgp=-0.01) - base_op),
            ("人件費 +10%", op_with(dH=+0.10) - base_op),
            ("人件費 -10%", op_with(dH=-0.10) - base_op),
            ("経費 +10%", op_with(dK=+0.10) - base_op),
            ("経費 -10%", op_with(dK=-0.10) - base_op),
        ]
        df = pd.DataFrame(changes, columns=["ドライバ","OP変化（円）"])
        return df

    df_sens = recompute_sensitivity_table()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheets_written = 0
        if isinstance(df_num, pd.DataFrame) and not df_num.empty:
            df_num.to_excel(writer, sheet_name="金額", index=True)
            sheets_written += 1
        if isinstance(df_kpi, pd.DataFrame) and not df_kpi.empty:
            df_kpi.to_excel(writer, sheet_name="KPI", index=True)
            sheets_written += 1
        if isinstance(df_sens, pd.DataFrame) and not df_sens.empty:
            df_sens.to_excel(writer, sheet_name="感応度", index=False)
            sheets_written += 1
        if sheets_written == 0:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet1")

        wb = writer.book
        if "金額" in wb.sheetnames:
            ws = wb["金額"]
            format_money_and_percent(ws, list(range(2, ws.max_column + 1)), [])
        if "KPI" in wb.sheetnames:
            ws = wb["KPI"]
            money_fmt = "\"¥\"#,##0;[Red]-\"¥\"#,##0"
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == "労働分配率":
                    for c in range(2, ws.max_column + 1):
                        ws.cell(row=r, column=c).number_format = "0.0%"
                else:
                    for c in range(2, ws.max_column + 1):
                        ws.cell(row=r, column=c).number_format = money_fmt
        if "感応度" in wb.sheetnames:
            ws = wb["感応度"]
            format_money_and_percent(ws, [2], [])

        meta_ws = wb.create_sheet("メタ情報")
        meta_data = [
            ("作成日時", dt.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("会計年度", fiscal_year),
            ("表示単位", unit),
            ("FTE", fte),
            ("ベース売上（円）", base_sales),
        ]
        for i, (k, v) in enumerate(meta_data, start=1):
            meta_ws.cell(row=i, column=1, value=k)
            meta_ws.cell(row=i, column=2, value=v)
        format_money_and_percent(meta_ws, [2], [])

        apply_japanese_styles(wb)
    data = output.getvalue()

    st.download_button(
        label="📥 Excel（.xlsx）をダウンロード",
        data=data,
        file_name=f"利益計画_{dt.date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.caption("© 経営計画策定WEBアプリ（Streamlit版） | 表示単位と計算単位を分離し、丸めの影響を最小化しています。")
